import openpyxl
from openpyxl.styles import Font
file_name1 = 'list1.xlsx'
file_name2 = 'list2.xlsx'


indexs2 = ['A8', 'B8', 'C8', 'D8']
indexs1 = ['B6', 'C6', 'D6', 'E6', 'F6', 'G6', 'H6',
          'I6', 'J6', 'K6', 'L6', 'M6', 'N6']
indexs4 = ['D8', 'E8', 'F8', 'H8', 'I8', 'J8', 'K8']





def get_exel(data, indexs=indexs2):
        workbook = openpyxl.load_workbook("list2.xlsx")
        # Выбор активного листа
        sheet = workbook.active
        for index, value in enumerate(indexs):
                sheet[value].font = Font(name='Times New Roman', size=14)
                sheet[value] = data[index]
        workbook.save(f'Перечень APM ИЦ для присоединения к аттестату.xlsx')


def get_list1(data, indexs=indexs1):
        workbook = openpyxl.load_workbook("list1.xlsx")
        # Выбор активного листа
        sheet = workbook.active
        for index, value in enumerate(indexs):
                sheet[value].font = Font(name='Times New Roman', size=14)
                sheet[value] = data[index]
        workbook.save(f'final_list1.xlsx')


def get_exel4(data, indexs=indexs4):
        workbook = openpyxl.load_workbook("list4.xlsx")
        # Выбор активного листа
        sheet = workbook.active
        for index, value in enumerate(indexs):
                sheet[value].font = Font(name='Times New Roman', size=14)
                sheet[value] = data[index]
        workbook.save(f'Заявка на подключение АРМ к сети ViPNet УМВД России по Курганской области.xlsx')


def get_list3(data: list, file_name: str):
        workbook = openpyxl.load_workbook(file_name)
        # Выбор активного листа
        sheet = workbook.active
        sheet["A8"].font = Font(name='Times New Roman', size=14)
        sheet["B8"].font = Font(name='Times New Roman', size=14)
        sheet["C8"].font = Font(name='Times New Roman', size=14)
        sheet["D8"].font = Font(name='Times New Roman', size=14)
        sheet["A8"] = data[0]
        sheet["B8"] = data[1]
        sheet["C8"] = data[2]
        sheet["D8"] = data[3]
        workbook.save(f'final_{file_name}')



def change_style(file_name):
        lst = ['A', 'B', 'C', 'D']
        workbook = openpyxl.load_workbook(file_name)
        # Выбор активного листа
        sheet = workbook.active
        for i in lst:
                sheet[f"{i}7"].font = Font(name='Times New Roman', size=14)

        workbook.save(f'final_{file_name}')



# get_exel(data2)
# get_list11(data1)
# get_list3(data2, file_name2)
# get_exel(data1, file_name1, indexs1)
# get_exel(data2, file_name2, indexs2)

# change_style("final_list1.xlsx")
